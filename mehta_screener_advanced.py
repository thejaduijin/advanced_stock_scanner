#!/usr/bin/env python3
"""
Mehta 3/3 NSE Top-750 Screener — ADVANCED EDITION

Run:
    python mehta_screener_advanced.py

Outputs:
    output/Mehta_Screener_YYYY-MM-DD.xlsx

Advanced Logic:
- Pillar 1 (Price): ATH breakout + volume confirmation + pullback depth + base count + EMA trend
- Pillar 2 (Fundamentals): PAT CAGR + acceleration + OCF validation + record margin
- Pillar 3 (RS): Composite RS score vs Nifty500 + Sector + RS line slope + Alpha/Beta
- Risk Overlay: Liquidity, market cap, debt, promoter holding gates
"""

import json, logging, math, re, sys, time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
CFG = json.loads((BASE / "config.json").read_text(encoding="utf-8"))
OUT = BASE / CFG["output"]["directory"]
OUT.mkdir(parents=True, exist_ok=True)

logging.basicConfig(
    level=getattr(logging, CFG["runtime"].get("log_level", "INFO")),
    format="%(asctime)s | %(levelname)s | %(message)s"
)
log = logging.getLogger("mehta")
for _name in ("yfinance", "yfinance.multi", "yfinance.utils"):
    _yl = logging.getLogger(_name)
    _yl.setLevel(logging.CRITICAL)
    _yl.propagate = False

HEADERS = {
    "User-Agent": CFG["data"]["user_agent"],
    "Accept-Language": "en-US,en;q=0.9",
}

session = requests.Session()
session.headers.update(HEADERS)


# ── Config thresholds ─────────────────────────────────────────────────────────
RULES = CFG.get("rules", {})
PRICE_QUALITY_MIN = RULES.get("price_quality_min", 70)
FUNDAMENTAL_MIN = RULES.get("fundamental_min", 70)
RS_COMPOSITE_MIN = RULES.get("rs_composite_min", 75)
MIN_AVG_VOLUME = RULES.get("min_avg_volume", 100000)
MIN_MARKET_CAP_CR = RULES.get("min_market_cap_cr", 500)


def clean_symbol(s):
    s = str(s).strip().upper()
    return s.replace(".NS", "").replace(".BO", "")


def yahoo_symbol(s):
    raw = str(s).strip().upper()
    if raw.startswith("^"):
        return raw
    if raw.endswith(".NS") or raw.endswith(".BO"):
        return raw
    return clean_symbol(raw) + ".NS"


def get_json(url, params=None):
    r = session.get(url, params=params, timeout=CFG["runtime"]["http_timeout"])
    r.raise_for_status()
    return r.json()


def fetch_nse_universe():
    local = CFG["universe"].get("local_csv", "")
    if local:
        p = BASE / local
        if p.exists():
            try:
                df = pd.read_csv(p)
                normalized = normalize_universe(df)
                if len(normalized) >= CFG["universe"].get("minimum_universe_size", 1):
                    return normalized
                log.warning("Local universe %s too small (%d); trying remote.", p.name, len(normalized))
            except Exception as e:
                log.warning("Local universe unreadable: %s", e)

    urls = CFG["universe"]["urls"]
    for url in urls:
        try:
            log.info("Downloading universe: %s", url)
            r = session.get(url, timeout=min(8, int(CFG["runtime"].get("http_timeout", 12))))
            r.raise_for_status()
            if url.lower().endswith(".csv"):
                from io import StringIO
                df = pd.read_csv(StringIO(r.text))
            else:
                tables = pd.read_html(r.text)
                df = max(tables, key=len)
            df = normalize_universe(df)
            if len(df) >= CFG["universe"]["minimum_universe_size"]:
                return df
        except Exception as e:
            log.warning("Universe source failed: %s", e)

    raise RuntimeError("Could not obtain a sufficiently large NSE universe.")


def normalize_universe(df):
    cols = {str(c).strip().lower(): c for c in df.columns}
    sym_col = next((cols[k] for k in ["symbol", "ticker", "code"] if k in cols), None)
    if sym_col is None:
        raise ValueError("Universe file needs a Symbol/Ticker/Code column.")

    out = pd.DataFrame()
    out["Symbol"] = df[sym_col].map(clean_symbol)
    for target, candidates in {
        "Company": ["company name", "company", "name"],
        "Industry": ["industry", "industry name", "sector"],
    }.items():
        c = next((cols[k] for k in candidates if k in cols), None)
        out[target] = df[c].astype(str) if c else ""

    # Capture Market Cap if present
    mc_col = next((cols[k] for k in ["market cap", "marketcap", "mcap"] if k in cols), None)
    out["Market Cap"] = df[mc_col].astype(float) if mc_col else 0.0

    out = out.drop_duplicates("Symbol")
    out = out[out["Symbol"].str.match(r"^[A-Z0-9&.-]+$", na=False)]
    return out.reset_index(drop=True)


def yf_download(symbols, period="2y"):
    tickers = [yahoo_symbol(s) for s in symbols if clean_symbol(s)]
    if not tickers:
        raise RuntimeError("The NSE universe is empty.")

    frames = []
    batch_size = int(CFG["runtime"].get("yfinance_batch_size", 100))
    retries = int(CFG["runtime"].get("yfinance_retries", 2))
    timeout = int(CFG["runtime"].get("http_timeout", 12))

    for start in range(0, len(tickers), batch_size):
        batch = tickers[start:start + batch_size]
        last_error = None
        for attempt in range(1, retries + 1):
            try:
                log.info("Downloading prices %d-%d/%d (attempt %d)",
                         start + 1, min(start + len(batch), len(tickers)),
                         len(tickers), attempt)
                data = yf.download(
                    tickers=batch,
                    period=period,
                    interval="1d",
                    auto_adjust=False,
                    progress=False,
                    group_by="ticker",
                    threads=True,
                    timeout=timeout,
                )
                if data is not None and not data.empty:
                    frames.append(data)
                    last_error = None
                    break
                last_error = RuntimeError("Yahoo returned no price rows")
            except Exception as e:
                last_error = e
                time.sleep(min(2 * attempt, 6))
        if last_error is not None:
            log.warning("Price batch %d-%d failed after %d attempts: %s",
                        start + 1, start + len(batch), retries, last_error)

    if not frames:
        raise RuntimeError("Yahoo Finance returned no price data for any ticker.")

    return pd.concat(frames, axis=1)


def extract_close(data, symbol):
    y = yahoo_symbol(symbol)
    if isinstance(data.columns, pd.MultiIndex):
        if y not in data.columns.get_level_values(0):
            return pd.Series(dtype=float)
        x = data[y]
        return x["Close"].dropna() if "Close" in x.columns else pd.Series(dtype=float)
    return data["Close"].dropna() if "Close" in data.columns else pd.Series(dtype=float)


def extract_volume(data, symbol):
    y = yahoo_symbol(symbol)
    if isinstance(data.columns, pd.MultiIndex):
        if y not in data.columns.get_level_values(0):
            return pd.Series(dtype=float)
        x = data[y]
        return x["Volume"].dropna() if "Volume" in x.columns else pd.Series(dtype=float)
    return data["Volume"].dropna() if "Volume" in data.columns else pd.Series(dtype=float)


def safe_return(series, days=252):
    if len(series) < days + 1:
        return np.nan
    return float(series.iloc[-1] / series.iloc[-days-1] - 1)


# ═══════════════════════════════════════════════════════════════════════════════
# ADVANCED PILLAR 1: PRICE QUALITY
# ═══════════════════════════════════════════════════════════════════════════════

def price_quality_score(close: pd.Series, volume: pd.Series) -> dict:
    if close.empty or len(close) < 60:
        return {"Price Quality Score": 0.0, "Is Strong Setup": False}

    current = float(close.iloc[-1])
    ath = float(close.max())
    ath_score = min(current / ath, 1.0) * 100

    # Pullback depth from recent high (last 20 days)
    recent = close.iloc[-20:]
    recent_high = float(recent.max())
    recent_low = float(recent.min())
    dip = (recent_high - recent_low) / recent_high if recent_high > 0 else 1.0
    dip_score = max(0, 100 - dip * 200)

    # Volume confirmation
    vol_score = 0.0
    if not volume.empty and len(volume) >= 20:
        avg_vol = float(volume.iloc[-20:].mean())
        recent_vol = float(volume.iloc[-5:].mean())
        vol_score = min(recent_vol / avg_vol, 3.0) / 3.0 * 100 if avg_vol > 0 else 0.0

    # Time consolidating near ATH (last 60 days)
    near_ath = close >= ath * 0.95
    base_weeks = float(near_ath.iloc[-60:].sum()) / 5.0
    base_score = min(base_weeks / 4.0, 1.0) * 100

    # 200 EMA trend
    ema200 = close.ewm(span=200, adjust=False).mean()
    ema_trend = float(ema200.iloc[-1] / ema200.iloc[-20] - 1) * 100 if len(ema200) >= 20 else 0.0
    trend_score = min(max(ema_trend * 10.0, 0.0), 100.0)

    total = ath_score * 0.25 + dip_score * 0.25 + vol_score * 0.20 + base_score * 0.15 + trend_score * 0.15

    return {
        "Price Quality Score": round(total, 1),
        "Is Strong Setup": total >= PRICE_QUALITY_MIN,
        "ATH Distance %": round(current / ath - 1, 4),
        "Max Drawdown from High": round(dip, 4),
        "Volume Ratio": round(recent_vol / avg_vol, 2) if not volume.empty else None,
        "Weeks Near ATH": round(base_weeks, 1),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ADVANCED PILLAR 2: FUNDAMENTAL QUALITY
# ═══════════════════════════════════════════════════════════════════════════════

def fundamental_score(pat_history: list, ocf_history: list = None) -> dict:
    """
    pat_history: list of annual PAT values (oldest to newest), last = TTM
    ocf_history: optional list of OCF values aligned with pat_history
    """
    if not pat_history or len(pat_history) < 3:
        return {"Fundamental Score": 0.0, "Is Quality": False}

    # Clean data
    clean_pat = [float(p) for p in pat_history if p is not None and not (isinstance(p, float) and math.isnan(p))]

    if len(clean_pat) < 3:
        return {"Fundamental Score": 0.0, "Is Quality": False}

    # 1. Growth Consistency
    first = abs(clean_pat[0]) if clean_pat[0] != 0 else 1.0
    cagr = (clean_pat[-1] / first) ** (1.0 / len(clean_pat)) - 1
    negative_years = sum(1 for p in clean_pat if p < 0)
    consistency_score = max(0, 100 - negative_years * 30) * (1.0 if cagr > 0.10 else 0.5)

    # 2. Acceleration
    yoy = []
    for i in range(1, len(clean_pat)):
        prev = abs(clean_pat[i-1]) if clean_pat[i-1] != 0 else 1.0
        yoy.append((clean_pat[i] - clean_pat[i-1]) / prev)
    acceleration = yoy[-1] - yoy[-2] if len(yoy) >= 2 else 0.0
    accel_score = max(0, min(100, 50 + acceleration * 100))

    # 3. Cash Flow Quality
    cash_quality = 100
    if ocf_history and len(ocf_history) >= len(clean_pat):
        for pat, ocf in zip(clean_pat, ocf_history):
            pat_abs = abs(pat) if pat != 0 else 1.0
            if ocf is not None and not math.isnan(ocf) and ocf < pat_abs * 0.7:
                cash_quality -= 20

    # 4. Record PAT margin
    historical = clean_pat[:-1] if len(clean_pat) > 1 else clean_pat
    record = max(historical) if historical else clean_pat[-1]
    ttm = clean_pat[-1]
    margin = (ttm - record) / abs(record) if record != 0 else 0.0
    record_score = 100 if margin > 0.05 else 70 if margin > 0 else 0

    total = consistency_score * 0.25 + accel_score * 0.25 + cash_quality * 0.25 + record_score * 0.25

    return {
        "Fundamental Score": round(total, 1),
        "Is Quality": total >= FUNDAMENTAL_MIN and record_score > 0,
        "PAT CAGR": round(cagr, 4),
        "Earnings Acceleration": round(acceleration, 4),
        "OCF Quality": cash_quality,
        "Record PAT Margin": round(margin, 4),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ADVANCED PILLAR 3: RELATIVE STRENGTH
# ═══════════════════════════════════════════════════════════════════════════════

def relative_strength_metrics(close: pd.Series, bench_close: pd.Series, sector_close: pd.Series) -> dict:
    if close.empty or bench_close.empty or len(close) < 60 or len(bench_close) < 60:
        return {"RS Composite Score": 0, "Is RS Leader": False}

    lookback = min(252, len(close), len(bench_close), len(sector_close))
    if lookback < 30:
        return {"RS Composite Score": 0, "Is RS Leader": False}

    stock_52w = close.iloc[-1] / close.iloc[-lookback] - 1
    nifty_52w = bench_close.iloc[-1] / bench_close.iloc[-lookback] - 1
    sector_52w = sector_close.iloc[-1] / sector_close.iloc[-lookback] - 1

    # RS Line (stock / benchmark)
    rs_line = close / bench_close
    rs_sma50 = rs_line.rolling(50).mean()
    rs_sma200 = rs_line.rolling(200).mean()

    rs_trend = "UP"
    if len(rs_sma50) >= 1 and len(rs_sma200) >= 1:
        rs_trend = "UP" if rs_sma50.iloc[-1] > rs_sma200.iloc[-1] else "DOWN"

    rs_slope = (rs_line.iloc[-1] / rs_line.iloc[-min(20, len(rs_line)-1)] - 1) * 100 if len(rs_line) > 20 else 0.0

    # Alpha/Beta
    stock_daily = close.pct_change().dropna()
    bench_daily = bench_close.pct_change().dropna()
    aligned = pd.concat([stock_daily, bench_daily], axis=1).dropna()
    alpha, beta = 0.0, 1.0
    if len(aligned) > 30:
        cov = aligned.cov().iloc[0, 1]
        bvar = aligned.iloc[:, 1].var()
        beta = float(cov / bvar) if bvar != 0 else 1.0
        alpha = float((aligned.iloc[:, 0].mean() - beta * aligned.iloc[:, 1].mean()) * 252)

    beats_nifty = stock_52w > nifty_52w
    beats_sector = stock_52w > sector_52w
    rs_line_strong = rs_trend == "UP" and rs_slope > 0

    score = 0
    if beats_nifty: score += 25
    if beats_sector: score += 25
    if rs_line_strong: score += 25
    if alpha > 0.05: score += 25

    return {
        "RS Composite Score": score,
        "Is RS Leader": score >= RS_COMPOSITE_MIN,
        "52W vs Nifty": round(float(stock_52w - nifty_52w), 4),
        "52W vs Sector": round(float(stock_52w - sector_52w), 4),
        "RS Line Trend": rs_trend,
        "RS Momentum 20D": round(float(rs_slope), 4),
        "Beta": round(float(beta), 2),
        "Alpha (Annual)": round(float(alpha), 4),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# RISK OVERLAY
# ═══════════════════════════════════════════════════════════════════════════════

def risk_overlay(market_cap: float, avg_volume: float) -> dict:
    failures = []

    if avg_volume < MIN_AVG_VOLUME:
        failures.append(f"Illiquid (vol {avg_volume:,.0f} < {MIN_AVG_VOLUME:,.0f})")

    if market_cap > 0 and market_cap < MIN_MARKET_CAP_CR:
        failures.append(f"Too Small (MCap ₹{market_cap:,.0f} Cr < {MIN_MARKET_CAP_CR})")

    return {
        "Pass Risk Filter": len(failures) == 0,
        "Risk Flags": "; ".join(failures) if failures else "None",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# PAT FETCHING (enhanced to return history)
# ═══════════════════════════════════════════════════════════════════════════════

def screener_pat(symbol):
    url = f"https://www.screener.in/company/{clean_symbol(symbol)}/consolidated/"
    try:
        r = session.get(url, timeout=min(8, int(CFG["runtime"].get("http_timeout", 12))))
        if r.status_code != 200:
            url = f"https://www.screener.in/company/{clean_symbol(symbol)}/"
            r = session.get(url, timeout=CFG["runtime"]["http_timeout"])
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "html.parser")

        section = None
        for h2 in soup.find_all(["h2", "h3"]):
            if "Profit & Loss" in h2.get_text(" ", strip=True):
                section = h2.find_parent("section")
                break
        if section is None:
            return {}

        table = section.find("table")
        if table is None:
            return {}

        headers = [x.get_text(" ", strip=True) for x in table.find_all("th")]
        rows = []
        for tr in table.find_all("tr"):
            cells = [x.get_text(" ", strip=True) for x in tr.find_all(["th", "td"])]
            if cells:
                rows.append(cells)

        pat_row = None
        ocf_row = None
        for row in rows:
            label = row[0].lower()
            if label in {"net profit", "profit after tax", "net profit / loss"}:
                pat_row = row
            if label in {"cash from operating activity", "operating cash flow", "cash flow from operations"}:
                ocf_row = row

        def num(x):
            x = str(x).replace(",", "").replace("₹", "").strip()
            if x in {"", "-", "—", "nan"}:
                return np.nan
            m = re.search(r"-?\d+(?:\.\d+)?", x)
            return float(m.group()) if m else np.nan

        def extract_history(row_data):
            if not row_data:
                return []
            vals = [num(x) for x in row_data[1:]]
            return [v for v in vals if not math.isnan(v)]

        pat_history = extract_history(pat_row)
        ocf_history = extract_history(ocf_row)

        if not pat_history:
            return {}

        ttm = pat_history[-1]
        annual = pat_history[:-1] if len(pat_history) > 1 else pat_history
        record = max(annual) if annual else ttm

        return {
            "TTM PAT": ttm,
            "Record PAT": record,
            "Record PAT?": bool(ttm >= record * CFG["rules"]["pat_record_tolerance"]),
            "PAT Data Source": "Screener.in",
            "PAT History": pat_history,
            "OCF History": ocf_history,
            "Exceptional Adjustment": "NOT AUTOMATICALLY EXCLUDED",
        }
    except Exception as e:
        log.debug("PAT failed %s: %s", symbol, e)
        return {}


# ═══════════════════════════════════════════════════════════════════════════════
# BENCHMARKS
# ═══════════════════════════════════════════════════════════════════════════════

def sector_benchmark(industry):
    mapping = CFG["sector_benchmarks"]
    text = str(industry).lower()
    for key, ticker in mapping.items():
        if key.lower() in text:
            return ticker
    return CFG["default_sector_benchmark"]


def benchmark_returns():
    tickers = [CFG["benchmarks"]["nifty500"], CFG["default_sector_benchmark"]]
    tickers += list(CFG["sector_benchmarks"].values())
    tickers = list(dict.fromkeys(tickers))
    out = {}
    timeout = int(CFG["runtime"].get("benchmark_timeout", 8))

    for t in tickers:
        yahoo_ticker = yahoo_symbol(t)
        try:
            log.info("Downloading benchmark %s", yahoo_ticker)
            raw = yf.download(
                tickers=yahoo_ticker, period="2y", interval="1d",
                auto_adjust=False, progress=False, group_by="ticker",
                threads=False, timeout=timeout,
            )
            if raw is None or raw.empty:
                continue
            close = extract_close(raw, t)
            value = safe_return(close, 252)
            if np.isfinite(value):
                out[t] = close  # Store full series for advanced RS calc
        except Exception:
            log.warning("Benchmark %s unavailable; skipping", yahoo_ticker)

    return out


# ═══════════════════════════════════════════════════════════════════════════════
# SCORING ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def advanced_score_row(row: dict) -> dict:
    p1_pass = bool(row.get("Is Strong Setup", False))
    p2_pass = bool(row.get("Is Quality", False))
    p3_pass = bool(row.get("Is RS Leader", False))
    risk_pass = bool(row.get("Pass Risk Filter", True))

    score = int(p1_pass) + int(p2_pass) + int(p3_pass)

    # Weighted score (0.0 – 3.0)
    p1 = float(row.get("Price Quality Score", 0)) / 100.0
    p2 = float(row.get("Fundamental Score", 0)) / 100.0
    p3 = float(row.get("RS Composite Score", 0)) / 100.0
    weighted = p1 * 1.0 + p2 * 1.0 + p3 * 1.0  # Equal pillar weights

    if not risk_pass:
        action = "RISK REJECT"
    elif score == 3 and weighted >= 2.4:
        action = "CONVICTION BUY"
    elif score == 3:
        action = "SUPER PERFORMER - BUY"
    elif score == 2:
        action = "PERFORMER - HOLD"
    elif score == 1:
        action = "WEAK - WATCHLIST"
    else:
        action = "EXIT"

    failures = []
    if not p1_pass:
        failures.append("Price setup weak")
    if not p2_pass:
        failures.append("Fundamentals weak")
    if not p3_pass:
        failures.append("RS weak")
    if not risk_pass:
        failures.append(row.get("Risk Flags", ""))

    row["Score"] = f"{score}/3"
    row["Weighted Score"] = round(weighted, 2)
    row["Action"] = action
    row["Pass"] = risk_pass and score >= 2
    row["Failure Reason"] = "; ".join(f for f in failures if f)
    return row


# ═══════════════════════════════════════════════════════════════════════════════
# PROCESS STOCK (the main pipeline)
# ═══════════════════════════════════════════════════════════════════════════════

def process_stock(symbol, industry, company, prices, bench_series, market_cap=0.0):
    close = extract_close(prices, symbol)
    volume = extract_volume(prices, symbol)

    if close.empty:
        return {"Symbol": symbol, "Company": company, "Industry": industry,
                "Data Status": "No price data"}

    # Pillar 1: Price Quality
    pm = get_price_metrics(close)
    pq = price_quality_score(close, volume)

    # Pillar 2: Fundamentals
    pat_data = screener_pat(symbol)
    pat_history = pat_data.pop("PAT History", [])
    ocf_history = pat_data.pop("OCF History", [])
    fq = fundamental_score(pat_history, ocf_history)

    # Pillar 3: Relative Strength
    nifty_ticker = CFG["benchmarks"]["nifty500"]
    sector_ticker = sector_benchmark(industry)
    bench_close = bench_series.get(nifty_ticker, pd.Series(dtype=float))
    sector_close = bench_series.get(sector_ticker, bench_close)  # Fallback to nifty
    rs = relative_strength_metrics(close, bench_close, sector_close)

    # Risk Overlay
    avg_vol = float(volume.mean()) if not volume.empty else 0.0
    risk = risk_overlay(market_cap=market_cap, avg_volume=avg_vol)

    # Base metrics
    nifty_ret = safe_return(bench_close, 252) if not bench_close.empty else np.nan
    sec_ret = safe_return(sector_close, 252) if not sector_close.empty else np.nan
    stock_ret = safe_return(close, 252)

    row = {
        "Symbol": symbol,
        "Company": company,
        "Industry": industry,
        "Sector Benchmark": sector_ticker,
        **pm,
        **pat_data,
        **pq,
        **fq,
        **rs,
        **risk,
        "Nifty500 52W Return": nifty_ret,
        "Sector 52W Return": sec_ret,
        "52W Return": stock_ret,
        "Beats Nifty500": bool(np.isfinite(stock_ret) and np.isfinite(nifty_ret) and stock_ret > nifty_ret),
        "Beats Sector": bool(np.isfinite(stock_ret) and np.isfinite(sec_ret) and stock_ret > sec_ret),
        "Data Status": "Price OK",
    }

    # PAT is slow — skip if already disqualified by price + RS
    pre_score = int(pq.get("Is Strong Setup", False)) + int(rs.get("Is RS Leader", False))
    if pre_score == 0 and not pat_data.get("PAT Data Source"):
        row["PAT Data Source"] = "Skipped (not a 3/3 candidate)"

    row.setdefault("Record PAT?", False)
    return advanced_score_row(row)


def get_price_metrics(close):
    if close.empty:
        return {}
    current = float(close.iloc[-1])
    ath = float(close.max())
    ema200 = float(close.ewm(span=200, adjust=False).mean().iloc[-1])
    r52 = safe_return(close, 252)
    return {
        "Current Price": current,
        "ATH": ath,
        "ATH % From High": current / ath - 1,
        "At/Near ATH": current >= ath * CFG["rules"]["ath_tolerance"],
        "200 EMA": ema200,
        "Below 200 EMA": current < ema200,
        "52W Return": r52,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

def export_excel(df, path):
    sheets = {
        "ALL": df,
        "CONVICTION": df[df["Action"] == "CONVICTION BUY"],
        "SUPER 3-3": df[df["Score"] == "3/3"],
        "PERFORMER 2-3": df[df["Score"] == "2/3"],
        "RISK REJECTS": df[df["Action"] == "RISK REJECT"],
        "EXIT": df[df["Score"].isin(["0/3", "1/3"])],
    }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, x in sheets.items():
            if not x.empty:
                x.to_excel(writer, sheet_name=name[:31], index=False)

        summary = pd.DataFrame([
            ["Run date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Universe size", len(df)],
            ["Conviction (3/3 High Wt)", int((df["Action"] == "CONVICTION BUY").sum())],
            ["3/3 Super Performers", int((df["Score"] == "3/3").sum())],
            ["2/3 Performers", int((df["Score"] == "2/3").sum())],
            ["1/3 Weak", int((df["Score"] == "1/3").sum())],
            ["0/3 Exit", int((df["Score"] == "0/3").sum())],
            ["Risk Rejects", int((df["Action"] == "RISK REJECT").sum())],
        ], columns=["Metric", "Value"])
        summary.to_excel(writer, sheet_name="SUMMARY", index=False)

    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            width = min(max(max(len(str(c.value or "")) for c in col) + 2, 10), 35)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width
    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    start = time.time()
    log.info("Starting Mehta 3/3 NSE Screener (ADVANCED EDITION)")

    universe = fetch_nse_universe()
    if "Market Cap" in universe.columns:
        universe = universe.sort_values("Market Cap", ascending=False)
    universe = universe.head(int(CFG["universe"]["top_n"])).copy()

    symbols = universe["Symbol"].tolist()
    log.info("Screening %d symbols", len(symbols))

    prices = yf_download(symbols, period="2y")
    bench_series = benchmark_returns()

    results = []
    workers = int(CFG["runtime"].get("workers", 16))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {
            ex.submit(
                process_stock,
                r["Symbol"], r["Industry"], r["Company"], prices, bench_series,
                r.get("Market Cap", 0.0)
            ): r["Symbol"]
            for _, r in universe.iterrows()
        }
        for i, fut in enumerate(as_completed(futs), 1):
            try:
                results.append(fut.result())
            except Exception as e:
                sym = futs[fut]
                log.warning("%s failed: %s", sym, e)
                results.append({"Symbol": sym, "Data Status": f"ERROR: {e}"})
            if i % 50 == 0:
                log.info("Processed %d/%d", i, len(futs))

    df = pd.DataFrame(results)
    if df.empty:
        raise RuntimeError("No results generated.")

    score_order = {"3/3": 0, "2/3": 1, "1/3": 2, "0/3": 3, "RISK REJECT": 4}
    df["_sort"] = df["Score"].map(score_order).fillna(9)
    df = df.sort_values(["_sort", "Weighted Score"], ascending=[True, False]).drop(columns="_sort")

    date = datetime.now().strftime("%Y-%m-%d")
    path = OUT / f"Mehta_Screener_{date}.xlsx"
    export_excel(df, path)

    log.info("Done: %s", path)
    log.info("Elapsed: %.1fs", time.time() - start)

    print("\n" + "="*60)
    print("MEHTA 3/3 ADVANCED SCREEN COMPLETE")
    print("="*60)
    print(f"Stocks screened: {len(df)}")
    print(f"Conviction Buy: {(df['Action'] == 'CONVICTION BUY').sum()}")
    print(f"3/3 Super:      {(df['Score'] == '3/3').sum()}")
    print(f"2/3 Hold:       {(df['Score'] == '2/3').sum()}")
    print(f"1/3 Weak:       {(df['Score'] == '1/3').sum()}")
    print(f"0/3 Exit:       {(df['Score'] == '0/3').sum()}")
    print(f"Risk Rejects:   {(df['Action'] == 'RISK REJECT').sum()}")
    print(f"Excel: {path}")
    print("="*60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted.")
        sys.exit(130)
    except Exception as e:
        log.exception("Fatal error: %s", e)
        sys.exit(1)